import pandas as pd
from openpyxl import load_workbook
import os
import re


def extract_domain(text):
    """
    命名規則: 任意の接頭辞_ドメイン_接尾辞
    末尾からアンダーバーを2回辿って、真ん中（=ドメイン）を抽出する。
    接尾辞は固定でないが、必ずアンダーバーで始まる。
    例:
      for-ocr_www.alen.co.jp_product.html          -> www.alen.co.jp
      transcripted_for-ocr_www.alen.co.jp_product.html.xlsx -> www.alen.co.jp
    """
    name = os.path.basename(text)  # フォルダ名/ファイル名どちらでもOK
    last_us = name.rfind("_")
    if last_us == -1:
        return None
    left = name[:last_us]
    prev_us = left.rfind("_")
    if prev_us == -1:
        return None
    return left[prev_us + 1:]

def collect_transcripted_files(base_folder):
    target_folders = [
        os.path.join(base_folder, d)
        for d in os.listdir(base_folder)
        if os.path.isdir(os.path.join(base_folder, d)) and d.startswith("for-ocr_")
    ]
    return [
        os.path.join(folder, f)
        for folder in target_folders
        for f in os.listdir(folder)
        if f.startswith("transcripted_") and f.endswith(".xlsx")
    ]

def get_column_index_by_domain(ws, url_row_index, domain):
    for idx, cell in enumerate(ws[url_row_index][4:], start=5):
        if cell.value and domain in str(cell.value):
            return idx
    return None

def write_values_by_prefix(ws, rows_b, prefix, values, key_col_index):
    values_iter = iter(values)
    for r_idx, r_name in enumerate(rows_b, start=1):
        if r_name and r_name.startswith(prefix):
            if ws.cell(row=r_idx, column=key_col_index).value in [None, ""]:
                try:
                    ws.cell(row=r_idx, column=key_col_index).value = next(values_iter)
                except StopIteration:
                    break

def process_file(file_path, ws, rows_b, url_row_index):
    print(f"処理対象: {file_path}")
    domain = extract_domain(os.path.basename(file_path))
    if not domain:
        print(f"ドメインが抽出できませんでした: {file_path}")
        return

    if domain not in pd.ExcelFile(file_path).sheet_names:
        print(f"シート '{domain}' がファイルに存在しません。スキップします。")
        return

    df = pd.read_excel(file_path, sheet_name=domain, header=None, names=["row", "value"])
    col_idx = get_column_index_by_domain(ws, url_row_index, domain)
    if col_idx is None:
        print(f"警告: {domain} に対応するURL列が見つかりません。")
        return

    sub_strength = df[df["row"].str.startswith("subcategories_strength")]["value"].str.strip().str.lower()
    match_map = {str(r_name).strip().lower(): r_idx for r_idx, r_name in enumerate(rows_b, start=1) if r_name}
    for value in sub_strength:
        r_idx = match_map.get(value)
        if r_idx:
            ws.cell(row=r_idx, column=col_idx).value = "●"

    write_values_by_prefix(ws, rows_b, "emotional_hashtags",
        df[df["row"].str.startswith("emotional_hashtags")].sort_values("row")["value"].tolist(), col_idx)
    write_values_by_prefix(ws, rows_b, "functional_hashtags",
        df[df["row"].str.startswith("functional_hashtags")].sort_values("row")["value"].tolist(), col_idx)

def main():
    base_folder = r"●●●●●\●●●●●\website-analysis-pipeline\playwright_output"
    file_b = r"●●●●●\●●●●●\website-analysis-pipeline\Structure_Beauty.xlsx"
    transcripted_files = collect_transcripted_files(base_folder)

    print(f"実行予定：対象ファイル数 {len(transcripted_files)}")
    if input("実行しますか？ (y/n): ").strip().lower() != "y":
        print("中止しました。")
        return

    if not transcripted_files:
        raise FileNotFoundError("transcripted_ファイルが見つかりません。")

    wb = load_workbook(file_b)
    ws = wb.active
    rows_b = [cell.value for cell in ws["D"]]
    url_row_index = rows_b.index("URL") + 1

    for file_path in transcripted_files:
        process_file(file_path, ws, rows_b, url_row_index)

    output_path = file_b.replace(".xlsx", "_updated.xlsx")
    wb.save(output_path)
    print("処理完了:", output_path)

if __name__ == "__main__":
    main()
