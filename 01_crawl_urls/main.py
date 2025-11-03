import requests
from requests.adapters import HTTPAdapter
from urllib3 import PoolManager
import ssl
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse, urldefrag
import time
import openpyxl

# --- TLSアダプタ定義（暗号強度を下げて互換性を確保） ---
class TLSAdapter(HTTPAdapter):
    def init_poolmanager(self, *args, **kwargs):
        context = ssl.create_default_context()
        context.set_ciphers('DEFAULT@SECLEVEL=1')  # 古いTLS対応
        kwargs['ssl_context'] = context
        return super().init_poolmanager(*args, **kwargs)

# セッションを作成しTLSアダプタを適用
session = requests.Session()
session.mount("https://", TLSAdapter())

def normalize_url(url):
    url, _ = urldefrag(url)
    if url.startswith("http://"):
        url = "https://" + url[len("http://"):]
    parsed = urlparse(url)
    if "." not in parsed.path.split("/")[-1]:
        if not url.endswith("/"):
            url += "/"
    return url

def is_html_page(url):
    exclude_ext = (".jpg", ".jpeg", ".png", ".gif", ".webp", ".svg",
                   ".pdf", ".zip", ".tar", ".gz", ".mp4", ".mp3", ".css", ".js")
    return not url.lower().endswith(exclude_ext)

def crawl(start_url):
    visited = set()
    domain = urlparse(start_url).netloc
    path_filter = urlparse(start_url).path

    def _crawl(url):
        url = normalize_url(url)
        if url in visited or not is_html_page(url):
            return
        visited.add(url)

        try:
            response = session.get(url, timeout=10)
        except Exception as e:
            print(f"Error: {url} - {e}")
            return

        if 'text/html' not in response.headers.get('Content-Type', ''):
            return

        soup = BeautifulSoup(response.text, "html.parser")
        for link in soup.find_all("a", href=True):
            next_url = normalize_url(urljoin(url, link['href']))
            parsed = urlparse(next_url)
            if parsed.netloc == domain and parsed.path.startswith(path_filter):
                _crawl(next_url)

        time.sleep(0.5)

    _crawl(start_url)
    return sorted(visited)

# ===== Excel処理 =====
xlsx_path = r"●●●●●\●●●●●\website-analysis-pipeline\beautifulsoup-url_list.xlsx"
wb = openpyxl.load_workbook(xlsx_path)
ws = wb.active

# 1行目をURLヘッダとして利用
header = [cell.value for cell in ws[1] if cell.value and cell.value.startswith("http")]
if not header:
    raise ValueError("1行目にTarget URLが見つかりません")

for col, start_url in enumerate(header, start=1):
    if start_url and start_url.startswith("http"):
        print(f"\n[{col}] {start_url} をクロール中...")
        urls = crawl(start_url)
        # 既存の1行目の下からURLを書き込み
        for i, u in enumerate(urls, start=2):
            ws.cell(row=i, column=col, value=u)
        print(f" → {len(urls)}件取得")

wb.save(xlsx_path)
print("\n処理完了しました。")
